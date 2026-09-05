"""
ptz_shared.py
=====================================================
Shared infrastructure used by patrol_track_main.py:

  - camera / ONVIF / patrol / zoom-stage / YOLO / tracking config
  - FrameGrabber          (threaded "always latest frame" RTSP reader)
  - PTZController         (ONVIF SOAP: AbsoluteMove for patrol,
                            ContinuousMove for visual-servo tracking)
  - TargetTracker         (ByteTrack ID lock-on logic)
  - SharedState           (mode + PTZ-velocity handoff between threads)
  - ptz_command_loop()    (background thread that actually sends
                            ContinuousMove/Stop while tracking)
  - small helpers (box overlap, PPE alert throttling)

This file has no "patrol logic" and no YOLO model loading in it -
that all lives in patrol_track_main.py. This file is just the
plumbing both the patrol state machine and the tracking state
machine share.
"""

import base64
import json
import os
import queue
import time
import datetime
import threading

import cv2
import requests
import xml.etree.ElementTree as ET

from flask import Flask, Response
from requests.auth import HTTPDigestAuth
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook, load_workbook

# =====================================================
# CAMERA SETTINGS
# =====================================================

CAMERA_IP = "10.1.68.45"
USERNAME = "admin"
PASSWORD = "Admin@123"

PROFILE_TOKEN = "profiletoken01"    

PTZ_URL = f"http://{CAMERA_IP}/onvif/ptz_service"
MEDIA_URL = f"http://{CAMERA_IP}/onvif/media_service"

# Known RTSP stream URL (credentials already embedded).
RTSP_URL = f"rtsp://{USERNAME}:{PASSWORD}@{CAMERA_IP}:554/unicaststream/1"

# =====================================================
# PATROL SETTINGS (absolute-move sweep)
# =====================================================

HOME_PAN = -0.899
LEFT_PAN = -0.600
RIGHT_PAN = 0.770

# Default/fallback tilt. Individual patrol stages now use
# TILT_STAGES below (indexed the same way as ZOOM_STAGES /
# SPEED_STAGES) so tilt can vary per stage. This constant is kept
# only as a fallback for any code path that doesn't have a stage
# index handy.
TILT = -0.75

PATROL_SPEED = 0.10
DWELL_TIME = 0.01

POSITION_TOLERANCE = 0.015

# =====================================================
# ZOOM / TILT / SPEED STAGE SETTINGS
# =====================================================
# Stage 0 is a special "wide-tilt" pass added in front of the
# original 3-stage sweep: same pan waypoints and same zoom as the
# old Stage 1 (0.00), but with its own tilt (0.25) instead of the
# standard patrol tilt (-0.75).
#
#   Stage 0 -> zoom 0.00, tilt  0.25   (new)
#   Stage 1 -> zoom 0.00, tilt -0.75   (was "Stage 1" before)
#   Stage 2 -> zoom 0.38, tilt -0.75   (was "Stage 2" before)
#   Stage 3 -> zoom 0.75, tilt -0.75   (was "Stage 3" before)
#
# ZOOM_STAGES, TILT_STAGES, and SPEED_STAGES must all stay the
# same length and are indexed together (index i -> that stage's
# zoom + tilt + speed).

ZOOM_STAGES = [0.00, 0.00, 0.38, 0.75]
TILT_STAGES = [0.25, -0.75, -0.75, -0.75]

# Explicit per-stage patrol pan/tilt speed. Previously this was
# derived from PATROL_SPEED * (SPEED_REDUCTION_FACTOR ** i); now
# each stage's speed is set directly so it can be tuned
# independently instead of through a single global reduction
# factor. (Values below match what the old 0.60 reduction factor
# produced, as a starting point.)
SPEED_STAGES = [
    0.20,    # Stage 0 -> zoom 0.00, tilt  0.25
    0.16,    # Stage 1 -> zoom 0.00, tilt -0.75
    0.09,    # Stage 2 -> zoom 0.38, tilt -0.75
    0.038,   # Stage 3 -> zoom 0.75, tilt -0.75
]

assert len(SPEED_STAGES) == len(ZOOM_STAGES) == len(TILT_STAGES), (
    "ZOOM_STAGES, TILT_STAGES, and SPEED_STAGES must all be the same length"
)

ZOOM_SPEED = 0.5

# How long (seconds) the camera stays parked at HOME while the
# zoom motor physically moves before any pan/tilt is allowed.
ZOOM_HOLD_TIME = 3.0

# Dedicated (slow) pan/tilt speed used ONLY for the tilt-only move
# between Stage 0 and Stage 1 (both directions: Stage 0 -> Stage 1
# during the forward sweep, and Stage 1 -> Stage 0 during the
# end-of-sweep reset). Previously this tilt move had no explicit
# speed (0.0), so it ran at the camera's default/fastest rate.
STAGE0_1_TILT_SPEED = 0.3

# How long (seconds) the camera dwells at Stage 1's position during
# the end-of-sweep reset, after zooming back to Stage 1's level and
# before tilting on up to Stage 0.
STAGE1_RESET_DWELL_TIME = 2.0

# =====================================================
# YOLO / DETECTION SETTINGS (shared by patrol-trigger AND tracking)
# =====================================================

YOLO_MODEL_PATH = "best.pt"
YOLO_CONF_THRESHOLD = 0.6
YOLO_IMG_SIZE = 480

# ByteTrack config shipped with ultralytics.
TRACKER_CONFIG = "bytetrack.yaml"

PERSON_CLASS_NAME = "person"

PROCESS_WIDTH = 960
PROCESS_HEIGHT = 540

YOLO_DEVICE = 0  # falls back to CPU automatically if CUDA isn't available

WINDOW_NAME = "Patrol + PPE Tracking - Live Feed"

# =====================================================
# PPE COMPLIANCE SETTINGS (used only while actively tracking someone)
# =====================================================

PPE_OVERLAP_THRESHOLD = 0.3
VIOLATION_ALERT_COOLDOWN = 5.0

# -----------------------------------------------------
# PPE ITEM MAP (used by the Excel compliance log)
# -----------------------------------------------------
# Maps every PPE-related class name your model can output to
# (item_column_name, status). status is "OK" for the "wearing it"
# class and "VIOLATION" for the "not wearing it" class.
#
# Matched against your model's actual class list:
#   {0: 'boots', 1: 'gloves', 2: 'hardhat', 3: 'no_boots',
#    4: 'no_gloves', 5: 'no_hardhat', 6: 'no_vest',
#    7: 'person', 8: 'vest'}
#
# The 4 item names on the right ("Hardhat", "Vest", "Gloves",
# "Boots") become the 4 status columns in the Excel log - keep
# those consistent with PPE_ITEMS below.

PPE_CLASS_STATUS_MAP = {
    "hardhat":    ("Hardhat", "OK"),
    "no_hardhat": ("Hardhat", "VIOLATION"),

    "vest":       ("Vest", "OK"),
    "no_vest":    ("Vest", "VIOLATION"),

    "gloves":     ("Gloves", "OK"),
    "no_gloves":  ("Gloves", "VIOLATION"),

    "boots":      ("Boots", "OK"),
    "no_boots":   ("Boots", "VIOLATION"),
}

# Column order for the PPE status columns in the Excel sheet.
PPE_ITEMS = ["Hardhat", "Vest", "Gloves", "Boots"]

# Lowercase-keyed lookup so class-name matching is case-insensitive
# (mirrors how PPE_VIOLATION_CLASSES matching worked previously).
PPE_CLASS_STATUS_MAP_LOWER = {
    k.lower(): v for k, v in PPE_CLASS_STATUS_MAP.items()
}

# -----------------------------------------------------
# EXCEL LOG SETTINGS
# -----------------------------------------------------

EXCEL_LOG_PATH = "ppe_log.xlsx"

# Minimum seconds between actual disk saves. The in-memory sheet
# is updated every frame a tracked person's PPE is (re)checked,
# but the file itself is only written to disk at most this often,
# so we don't hammer the disk at 20-30fps. Open the file in Excel
# and re-open / refresh it every few seconds to see it "live".
EXCEL_SAVE_INTERVAL = 1.0

# =====================================================
# VISUAL SERVO (PAN/TILT FOLLOW) SETTINGS - tracking mode only
# =====================================================

PAN_KP = 0.6
TILT_KP = 0.6

# Hard ceiling on tracking speed. The proportional control below
# still tapers speed down as the person approaches center (so it
# eases in/out near the deadband instead of a hard on/off), but it
# can never command faster than this - 0.1 keeps PTZ motion slow
# and smooth even when the person first appears far off-center.
MAX_PAN_SPEED = 0.1
MAX_TILT_SPEED = 0.1

TILT_SIGN = -1.0

DEADBAND = 0.06

LOST_TARGET_GRACE_FRAMES = 10

COMMAND_RATE_HZ = 10.0

# =====================================================
# MODE-SWITCH TIMING
# =====================================================
# How long (seconds) the camera stays in tracking mode once a
# person is locked on, before handing control back to patrol.

TRACK_DURATION = 5.0

# =====================================================
# TRACKING SELECTION / RE-TRACK COOLDOWN
# =====================================================
# Once a given ByteTrack ID's TRACK_DURATION window ends, that ID
# must not be picked again as the active PTZ-follow target until
# this many seconds have passed - even if they're still standing
# right in front of the camera. This is separate from
# PPE_LOG_COOLDOWN_SECONDS below (which only gates Excel rows):
# this one gates who the camera physically follows next.

TRACKING_COOLDOWN_SECONDS = 120.0  # 2 minutes

# =====================================================
# PPE LOG COOLDOWN / SCREENSHOT SETTINGS
# =====================================================
# Once a given ByteTrack ID has been logged to the PPE Excel
# sheet, it must not be logged again until this many seconds have
# passed - even though tracking/PTZ-following of that person
# keeps working normally during the cooldown window. This stops
# the same person being written to the sheet on every single
# tracking session if they walk past the camera repeatedly.

PPE_LOG_COOLDOWN_SECONDS = 300.0

# Folder (relative to the working directory) where a snapshot
# image is saved each time a PPE log row is written, so the row's
# "Screenshot Path" column has something to point to.

SCREENSHOT_DIR = "screenshots"
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

# =====================================================
# NEW: BACKEND PPE ALERT SETTINGS
# =====================================================
# The ONLY section you ever need to edit.
#
# BACKEND_BASE_URL     - Backend server base URL (no trailing slash)
# BACKEND_CAMERA_ID    - Integer camera FK in the backend DB (ask dev)
# BACKEND_SITE_ID      - Integer site FK in the backend DB (ask dev)
# BACKEND_AUTH_TOKEN   - JWT access token (short-lived, ~1 hr)
# BACKEND_REFRESH_TOKEN- JWT refresh token (use to get new access token)
# BACKEND_USERNAME /   - Login credentials used when both tokens expire
# BACKEND_PASSWORD       (ask backend developer)
# =====================================================

BACKEND_BASE_URL      = "http://siteaense.kct.ac.in:8000"
BACKEND_SERVER_URL    = f"{BACKEND_BASE_URL}/api/ai-alerts/"
BACKEND_REFRESH_URL   = f"{BACKEND_BASE_URL}/api/auth/token/refresh/"  # Django SimpleJWT standard
BACKEND_LOGIN_URL     = f"{BACKEND_BASE_URL}/api/auth/login/"           # confirm with dev

# Camera / site integer FK IDs in the backend database
BACKEND_CAMERA_ID     = 1    # integer - camera_id column in ai_alerts table
BACKEND_SITE_ID       = 1    # integer - site_id column in ai_alerts table

# String used to identify this camera in filenames / local logging only
CAMERA_ID             = "camera_01"

# --- JWT tokens (from backend developer / login response) ---
# The worker auto-refreshes using BACKEND_REFRESH_TOKEN when the
# access token expires (HTTP 401).  When the refresh token also
# expires, it falls back to BACKEND_USERNAME / BACKEND_PASSWORD.
BACKEND_AUTH_TOKEN    = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJ0b2tlbl90eXBlIjoiYWNjZXNzIiwiZXhwIjoxNzg4NjEzMzU3LCJpYXQiOjE3ODg2"
    "MDk3NTcsImp0aSI6ImFhOTNlZmZlNTY5MjRmZDZiNWNjN2UyYTBjM2MyYWM2IiwidXNlcl"
    "9pZCI6IjEifQ.0d0mckilhMD03z0Gm8eWhy9wjBR8W7Y4BuLv53xS27A"
)
BACKEND_REFRESH_TOKEN = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJ0b2tlbl90eXBlIjoicmVmcmVzaCIsImV4cCI6MTc4ODYxMTU1NywiaWF0IjoxNzg4Nj"
    "A5NzU3LCJqdGkiOiI3MjcxY2EzM2Y0MDM0Mzg1OWE1Y2ZiNzA5Zjk3Mzg0NCIsInVzZXJf"
    "aWQiOiIxIn0.CcplpWwYhGoWo8q8GZRV9k_zQv939PIhmR0QVbp7cYo"
)

# Fallback login credentials (used when both tokens expire).
# Ask backend dev for a dedicated Jetson service account.
BACKEND_USERNAME      = "admin"       # backend login username
BACKEND_PASSWORD      = "Admin@123"   # backend login password

# --- Timing / Queue ---
PPE_ALERT_COOLDOWN_SECONDS = 15.0   # global cooldown: one alert per 15 s max
BACKEND_REQUEST_TIMEOUT    = 5.0    # seconds before giving up on HTTP POST
MAX_ALERT_QUEUE_SIZE       = 3      # drop oldest if queue exceeds this
BACKEND_SNAPSHOT_DIR       = "backend_alert_snapshots"
os.makedirs(BACKEND_SNAPSHOT_DIR, exist_ok=True)

# =====================================================
# NEW: INTEGRATED MJPEG FEED SERVER PORT
# =====================================================
FEED_SERVER_PORT = 8080

# How often (seconds) blocking wait-loops re-check for an
# interrupt / re-check camera position. Smaller = camera stops
# "immediately" when a person is seen, at the cost of slightly
# more chatter.
LOOP_POLL_INTERVAL = 0.1

# Number of consecutive unchanged-position polls (at
# LOOP_POLL_INTERVAL spacing) before a pan/tilt move is
# considered stalled.
STALL_COUNT_THRESHOLD = 70


# =====================================================
# SHARED STATE BETWEEN PATROL THREAD, DETECTION/TRACKING
# LOOP (main thread), AND THE PTZ COMMAND THREAD
# =====================================================
# mode:
#   "PATROL"   -> patrol thread is free to drive absolute moves
#   "TRACKING" -> patrol thread is parked, ptz_command_loop is
#                 free to drive continuous-move visual servo
#
# person_trigger_event:
#   SET by the main thread the instant a person is seen while
#   mode == "PATROL". The patrol thread polls this during every
#   move / dwell / zoom-hold and, when it sees it set, stops the
#   camera immediately and waits for tracking to finish.
#
# tracking_done_event:
#   SET by the main thread once TRACK_DURATION has elapsed and
#   continuous-move has been stopped. Tells the patrol thread it
#   may now issue the "return HOME at current zoom stage" move
#   and continue patrolling.

class SharedState:

    def __init__(self):

        self._lock = threading.Lock()
        self._mode = "PATROL"

        self.person_trigger_event = threading.Event()
        self.tracking_done_event = threading.Event()

        self._pan_speed = 0.0
        self._tilt_speed = 0.0
        self._locked = False

        # Purely informational, for the on-screen overlay.
        self.status_text = "Patrol - Stage 0"

        # Which patrol stage (0-3) was interrupted to start the
        # current tracking session. Set by the patrol thread right
        # before it hands off to tracking; read by the main
        # detection/tracking loop so PPE log rows can record which
        # stage the person was spotted during.
        self._patrol_stage = 0

    def get_mode(self):
        with self._lock:
            return self._mode

    def set_mode(self, mode):
        with self._lock:
            self._mode = mode

    def set_desired_velocity(self, pan_speed, tilt_speed, locked):
        with self._lock:
            self._pan_speed = pan_speed
            self._tilt_speed = tilt_speed
            self._locked = locked

    def get_desired_velocity(self):
        with self._lock:
            return self._pan_speed, self._tilt_speed, self._locked

    def set_status(self, text):
        with self._lock:
            self.status_text = text

    def get_status(self):
        with self._lock:
            return self.status_text

    def set_patrol_stage(self, stage_number):
        with self._lock:
            self._patrol_stage = stage_number

    def get_patrol_stage(self):
        with self._lock:
            return self._patrol_stage


# =====================================================
# PPE VIOLATION ALERT THROTTLING
# =====================================================

_violation_alert_lock = threading.Lock()
_last_alert_time = {}  # (track_id, violation_name) -> timestamp


def maybe_alert_violation(track_id, violation_name):
    key = (track_id, violation_name)
    now = time.time()

    with _violation_alert_lock:
        last = _last_alert_time.get(key, 0.0)
        if now - last < VIOLATION_ALERT_COOLDOWN:
            return
        _last_alert_time[key] = now

    print(f"\n[PPE VIOLATION] Person {track_id}: {violation_name}")


# =====================================================
# COOLDOWN MANAGER (generic per-ID "not too soon again" gate)
# =====================================================
# Reused for two independent purposes, each with its own instance
# and its own cooldown_seconds:
#
#   1. PPE Excel log dedup (PPE_LOG_COOLDOWN_SECONDS / 300s) -
#      gates whether a Track ID's PPE check produces a new
#      spreadsheet row. Uses should_log(), an atomic check+mark.
#
#   2. Tracking re-selection (TRACKING_COOLDOWN_SECONDS / 120s) -
#      gates whether a Track ID may be picked again as the active
#      PTZ-follow target. Uses is_on_cooldown() (read-only check,
#      can be called many times per frame without side effects)
#      plus mark_now() (called once, exactly when a person's
#      TRACK_DURATION window ends, to start their cooldown).
#
# These are separate CooldownManager instances - a person going on
# the 2-minute tracking cooldown has no effect on the 5-minute
# Excel-log cooldown or vice versa.

class CooldownManager:

    def __init__(self, cooldown_seconds=PPE_LOG_COOLDOWN_SECONDS):
        self.cooldown_seconds = cooldown_seconds
        self._lock = threading.Lock()
        self._last_used = {}  # track_id -> last-used timestamp

    def is_on_cooldown(self, track_id):
        """Read-only check - does NOT start/reset the cooldown.
        Safe to call repeatedly (e.g. once per visible person, every
        frame) while deciding who's eligible to track next."""

        with self._lock:
            last = self._last_used.get(track_id)

        if last is None:
            return False

        return (time.time() - last) < self.cooldown_seconds

    def mark_now(self, track_id):
        """Explicitly starts/resets this track_id's cooldown window
        from right now."""

        with self._lock:
            self._last_used[track_id] = time.time()

    def should_log(self, track_id):
        """Atomic check+mark: returns True (and starts a fresh
        cooldown window) if this track_id is not currently on
        cooldown, False otherwise. Used by the PPE Excel logger,
        where every eligible check should also count as the use."""

        now = time.time()

        with self._lock:
            last = self._last_used.get(track_id, 0.0)

            if now - last >= self.cooldown_seconds:
                self._last_used[track_id] = now
                return True

            return False

    def seconds_remaining(self, track_id):
        """Informational only (e.g. for on-screen display)."""

        with self._lock:
            last = self._last_used.get(track_id)

        if last is None:
            return 0.0

        return max(0.0, self.cooldown_seconds - (time.time() - last))


# =====================================================
# PPE EXCEL LOGGER
# =====================================================
# APPENDS ONE ROW PER LOG EVENT (not one row per person). A "log
# event" is: a tracked person's PPE was checked, and that Track ID
# is not currently on cooldown (see CooldownManager above) - so in
# practice a given person produces a new row at most once every
# PPE_LOG_COOLDOWN_SECONDS, no matter how many tracking sessions
# they trigger in that window.
#
# Columns:
#   Timestamp | Track ID | Hardhat | Vest | Gloves | Boots |
#   PPE Status | Screenshot Path
#
# Each PPE item gets its OWN column so it's immediately visible
# which specific item a given Track ID was violated on (e.g. the
# "Hardhat" column shows "VIOLATION" on that person's row if they
# weren't wearing one, "OK" if they were, blank if that item
# wasn't seen at all for this event).
#
# The workbook is kept open in memory and only saved to disk at
# most once every EXCEL_SAVE_INTERVAL seconds (see above), so you
# can leave the file open in Excel and refresh periodically to
# watch it update without generating a disk write on every frame.

EXCEL_HEADERS = (
    ["Timestamp", "Track ID"] + PPE_ITEMS + ["PPE Status", "Screenshot Path"]
)


class PPEExcelLogger:

    def __init__(self, path=EXCEL_LOG_PATH):

        self.path = path
        self._lock = threading.Lock()
        self._last_save = 0.0

        self._load_or_create()

    def _load_or_create(self):

        if os.path.exists(self.path):
            try:
                self.wb = load_workbook(self.path)
                self.ws = self.wb.active
                print(f"PPE log: appending to existing file '{self.path}'.")
                return
            except Exception as exc:
                print(f"PPE log: existing file '{self.path}' was corrupt/invalid ({exc}). Creating fresh file.")

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "PPE Log"

        self.ws.append(EXCEL_HEADERS)

        for col in range(1, len(EXCEL_HEADERS) + 1):
            self.ws.column_dimensions[
                self.ws.cell(row=1, column=col).column_letter
            ].width = 16

        self.wb.save(self.path)
        print(f"PPE log: created new file '{self.path}'.")

    def log(self, track_id, item_status, screenshot_path=None):
        """Appends ONE new row for this log event.

        item_status: dict of {item_name: "OK"/"VIOLATION"} for
        whichever PPE items were detected against this person at
        the moment of logging (item_name is one of PPE_ITEMS -
        "Hardhat", "Vest", "Gloves", "Boots"). Each item gets its
        own column in the row, so it's obvious at a glance exactly
        which item(s) this Track ID was violated on. Items not
        detected at all for this event are left blank. Overall
        "PPE Status" is "Non-Compliant" if any item is a
        VIOLATION, else "Compliant".
        """

        if not item_status:
            return

        violations = [
            item for item, status in item_status.items()
            if status == "VIOLATION"
        ]
        overall_status = "Non-Compliant" if violations else "Compliant"

        timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with self._lock:

            item_columns = [item_status.get(item, "") for item in PPE_ITEMS]

            self.ws.append(
                [timestamp_str, track_id]
                + item_columns
                + [overall_status, screenshot_path or ""]
            )

            now_t = time.time()
            if now_t - self._last_save >= EXCEL_SAVE_INTERVAL:
                self._save_locked()
                self._last_save = now_t

    def _save_locked(self):
        try:
            self.wb.save(self.path)
        except Exception as exc:
            print(f"PPE log: save failed ({exc}). Is the file open/locked in Excel?")

    def close(self):
        """Force a final save - call this on shutdown so the last
        few updates (which may be sitting inside the throttle
        window) aren't lost."""
        with self._lock:
            self._save_locked()


# =====================================================
# NEW: INTEGRATED MJPEG FEED SERVER
# =====================================================
# Hosts both raw and YOLO-annotated MJPEG streams on FEED_SERVER_PORT
# (default 8080).
#
# Routes:
#   /feed                  -> raw camera frame
#   /feed/0                -> raw camera frame (alias)
#   /feed/<camera_id>      -> raw camera frame (alias)
#   /annotated_feed        -> YOLO-annotated frame (bounding boxes)
#   /feed/annotated        -> same as /annotated_feed
#
# Usage (from patrol_track_main.py):
#   feed_server = FeedServer()
#   feed_server.start()
#   # inside detection loop:
#   feed_server.update_raw(frame)
#   feed_server.update_annotated(annotated_frame)

class _FrameBroadcaster:
    """Thread-safe, condition-variable-driven MJPEG frame broadcaster.
    update() is non-blocking; generate() is a streaming generator."""

    def __init__(self, quality=80):
        self.params = [
            int(cv2.IMWRITE_JPEG_QUALITY), quality,
            int(cv2.IMWRITE_JPEG_OPTIMIZE), 1,
        ]
        self._lock = threading.Lock()
        self._cond = threading.Condition(self._lock)
        self._jpeg_bytes = None
        self._seq = 0

    def update(self, frame):
        """Encode frame to JPEG and notify waiting generate() clients."""
        if frame is None:
            return
        ok, buf = cv2.imencode(".jpg", frame, self.params)
        if ok:
            with self._cond:
                self._jpeg_bytes = buf.tobytes()
                self._seq += 1
                self._cond.notify_all()

    def generate(self):
        """MJPEG streaming generator (one yield per new frame)."""
        last_seq = 0
        while True:
            with self._cond:
                while self._seq == last_seq or self._jpeg_bytes is None:
                    if not self._cond.wait(timeout=0.1):
                        break
                data = self._jpeg_bytes
                last_seq = self._seq
            if data is not None:
                yield (
                    b"--frame\r\n"
                    b"Content-Type: image/jpeg\r\n"
                    b"Content-Length: " + str(len(data)).encode() + b"\r\n\r\n"
                    + data + b"\r\n"
                )


class FeedServer:
    """Lightweight Flask MJPEG server running in a daemon thread.

    Serves both a raw and an annotated MJPEG stream from the
    detection loop in patrol_track_main.py without blocking it.
    update_raw() and update_annotated() are non-blocking calls.
    """

    def __init__(self, port=FEED_SERVER_PORT):
        self.port = port
        self._raw_bc  = _FrameBroadcaster(quality=75)   # raw camera
        self._ann_bc  = _FrameBroadcaster(quality=80)   # YOLO annotated
        self._app     = self._build_app()
        self._thread  = None

    # --------------------------------------------------
    def update_raw(self, frame):
        """Call every frame with the raw (un-annotated) camera frame."""
        self._raw_bc.update(frame)

    def update_annotated(self, frame):
        """Call every frame with the YOLO-annotated (bounding-box) frame."""
        self._ann_bc.update(frame)

    # --------------------------------------------------
    def _mjpeg_response(self, broadcaster):
        res = Response(
            broadcaster.generate(),
            mimetype="multipart/x-mixed-replace; boundary=frame",
        )
        res.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        res.headers["Pragma"]        = "no-cache"
        res.headers["Expires"]       = "0"
        return res

    def _build_app(self):
        import logging
        log = logging.getLogger("werkzeug")
        log.setLevel(logging.ERROR)  # suppress Flask per-request logs

        app = Flask(__name__ + "_feed")

        raw_bc = self._raw_bc
        ann_bc = self._ann_bc
        mjpeg  = self._mjpeg_response

        @app.route("/feed")
        @app.route("/feed/0")
        @app.route("/feed/<path:camera_id>")
        def raw_feed(camera_id=None):
            return mjpeg(raw_bc)

        @app.route("/annotated_feed")
        @app.route("/feed/annotated")
        def annotated_feed():
            return mjpeg(ann_bc)

        @app.route("/")
        def index():
            return (
                "Jetson PPE Feed Server<br>"
                "<a href='/feed'>/feed</a> &mdash; raw camera<br>"
                "<a href='/annotated_feed'>/annotated_feed</a> "
                "&mdash; YOLO annotated"
            )

        return app

    # --------------------------------------------------
    def start(self):
        """Start the Flask server in a background daemon thread."""
        def _run():
            self._app.run(
                host="0.0.0.0",
                port=self.port,
                threaded=True,
                use_reloader=False,
            )

        self._thread = threading.Thread(
            target=_run, daemon=True, name="FeedServer"
        )
        self._thread.start()
        print(f"[FEED SERVER] Started on port {self.port}")
        print(f"[FEED SERVER]  Raw feed:       http://0.0.0.0:{self.port}/feed")
        print(f"[FEED SERVER]  Annotated feed: http://0.0.0.0:{self.port}/annotated_feed")

    def stop(self):
        """No-op: the thread is a daemon and exits with the process."""
        pass


# =====================================================
# BOX OVERLAP HELPER (attributing a PPE box to a tracked person)
# =====================================================

def intersection_over_box_area(inner_box, outer_box):
    """Fraction of inner_box's area that overlaps outer_box."""

    ix1 = max(inner_box[0], outer_box[0])
    iy1 = max(inner_box[1], outer_box[1])
    ix2 = min(inner_box[2], outer_box[2])
    iy2 = min(inner_box[3], outer_box[3])

    iw = max(0.0, ix2 - ix1)
    ih = max(0.0, iy2 - iy1)
    intersection = iw * ih

    inner_area = max(
        1e-6,
        (inner_box[2] - inner_box[0]) * (inner_box[3] - inner_box[1])
    )

    return intersection / inner_area


# =====================================================
# TARGET LOCK-ON LOGIC (ByteTrack ID persistence)
# =====================================================
# TargetTracker's only job is: "is the ID I'm currently locked onto
# still visible this frame?" - it does NOT decide which ID to lock
# onto in the first place. That decision (which person to track
# next, respecting the 2-minute tracking cooldown) is made
# explicitly by the caller via select_next_target() below, and then
# handed to this tracker by setting .target_id directly. Keeping
# these two responsibilities separate is what makes "only track
# using Track ID, never by re-detection/proximity" hold everywhere.

class TargetTracker:

    def __init__(self):
        self.target_id = None
        self.missing_frames = 0

    def update(self, persons):
        """persons: list of dicts with keys id, cx, cy, box - cx/cy are
        normalized center-x/y in [-1, 1] relative to frame center.

        Looks for self.target_id among this frame's persons. If
        found, returns it. If not found, allows up to
        LOST_TARGET_GRACE_FRAMES consecutive missed frames (e.g. a
        brief occlusion) before giving up and clearing target_id.
        Returns None whenever nothing is currently locked on, or the
        locked target isn't visible this frame."""

        if self.target_id is None:
            return None

        for p in persons:
            if p["id"] == self.target_id:
                self.missing_frames = 0
                return p

        self.missing_frames += 1
        if self.missing_frames > LOST_TARGET_GRACE_FRAMES:
            self.target_id = None
            self.missing_frames = 0

        return None


def select_next_target(persons, tracking_cooldown_mgr):
    """Picks who the camera should track next, from the people
    visible THIS frame - purely by Track ID, never by re-detection
    or proximity to center.

    Eligible = not currently on the 2-minute tracking cooldown
    (i.e. wasn't tracked in the last TRACKING_COOLDOWN_SECONDS).
    Among the eligible people, picks the SMALLEST Track ID: since
    ByteTrack assigns IDs in the order tracks are first created,
    the smallest ID still visible is the earliest-detected person
    still in frame - so this naturally tracks people in
    first-detected order, then moves on to the next one.

    Returns the person dict to track next, or None if nobody
    currently visible is eligible (everyone in frame was already
    tracked within the last 2 minutes, or the frame is empty) - the
    caller should resume/continue patrol in that case.
    """

    eligible = [
        p for p in persons
        if not tracking_cooldown_mgr.is_on_cooldown(p["id"])
    ]

    if not eligible:
        return None

    return min(eligible, key=lambda p: p["id"])


# =====================================================
# THREADED FRAME GRABBER
# =====================================================

class FrameGrabber:

    def __init__(self, rtsp_url):

        self.rtsp_url = rtsp_url
        self._init_ffmpeg_env()
        self.cap = self._open_capture()

        if not self.cap.isOpened():
            raise RuntimeError(
                "Could not open RTSP stream. Check the URI, "
                "credentials, and network access to the camera."
            )

        self.lock = threading.Lock()
        self.latest_frame = None
        self.running = False
        self.thread = None
        self._last_frame_time = time.time()

    def _init_ffmpeg_env(self):
        os.environ["OPENCV_FFMPEG_CAPTURE_OPTIONS"] = (
            "rtsp_transport;tcp|fflags;nobuffer|flags;low_delay|max_delay;500000|reorder_queue_size;0"
        )

    def _open_capture(self):
        self._init_ffmpeg_env()
        cap = cv2.VideoCapture(self.rtsp_url, cv2.CAP_FFMPEG)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        return cap

    def start(self):

        self.running = True

        def loop():
            fail_count = 0
            while self.running:
                ret, frame = self.cap.read()
                now = time.time()
                if not ret or frame is None:
                    fail_count += 1
                    time.sleep(0.04)
                    if (now - self._last_frame_time) > 2.0 or fail_count > 40:
                        print("\n[FrameGrabber] RTSP stream stalled. Reconnecting...")
                        try:
                            self.cap.release()
                        except Exception:
                            pass
                        time.sleep(0.3)
                        self.cap = self._open_capture()
                        self._last_frame_time = time.time()
                        fail_count = 0
                    continue

                fail_count = 0
                self._last_frame_time = now
                with self.lock:
                    self.latest_frame = frame

        self.thread = threading.Thread(target=loop, daemon=True, name="FrameGrabber")
        self.thread.start()
        return self

    def read(self):
        with self.lock:
            if self.latest_frame is None:
                return False, None
            return True, self.latest_frame.copy()

    def stop(self):
        self.running = False
        if self.thread:
            self.thread.join(timeout=1)
        if self.cap:
            try:
                self.cap.release()
            except Exception:
                pass


# =====================================================
# PTZ CONTROLLER
# =====================================================
# Supports BOTH:
#   - AbsoluteMove + polling  -> used by the PATROL state machine
#   - ContinuousMove          -> used by the TRACKING visual servo
# on the same ONVIF session, since only one of the two is ever
# actively commanding the camera at a given moment (see SharedState
# mode + the coordination in patrol_track_main.py).

class PTZController:

    def __init__(self):

        self._create_session()
        self.auth = HTTPDigestAuth(USERNAME, PASSWORD)

        self.pan = None
        self.tilt = None
        self.zoom = None

        self.running = False
        self.poll_thread = None

    def _create_session(self):
        self.session = requests.Session()
        retry = Retry(total=2, backoff_factor=0.2)
        adapter = HTTPAdapter(max_retries=retry, pool_connections=1, pool_maxsize=2)
        self.session.mount("http://", adapter)

    # =================================================
    # SOAP
    # =================================================

    def send_soap(self, url, body, quiet=False):

        soap = f"""<?xml version="1.0" encoding="UTF-8"?>
<s:Envelope xmlns:s="http://www.w3.org/2003/05/soap-envelope">
<s:Body>
{body}
</s:Body>
</s:Envelope>"""

        try:
            response = self.session.post(
                url,
                data=soap,
                headers={"Content-Type": "application/soap+xml"},
                auth=self.auth,
                timeout=2.5,
            )
        except requests.RequestException as exc:
            if not quiet:
                print(f"ONVIF request to {url} failed: {exc}")
            self._create_session()
            return None

        if not quiet and response is not None:

            if response.status_code >= 300:
                print(
                    f"ONVIF command to {url} failed: "
                    f"HTTP {response.status_code} - "
                    f"{response.text[:300]}"
                )

            elif "Fault" in response.text:
                print(
                    f"ONVIF command to {url} returned a SOAP "
                    f"Fault: {response.text[:300]}"
                )

        return response

    # =================================================
    # STATUS / POLLING (used by the patrol state machine)
    # =================================================

    def get_status(self):

        body = f"""
<tptz:GetStatus
xmlns:tptz="http://www.onvif.org/ver20/ptz/wsdl">
<tptz:ProfileToken>{PROFILE_TOKEN}</tptz:ProfileToken>
</tptz:GetStatus>
"""

        try:
            response = self.send_soap(PTZ_URL, body, quiet=True)
            if response is None:
                return

            root = ET.fromstring(response.text)

            for elem in root.iter():
                if elem.tag.endswith("PanTilt"):
                    self.pan = float(elem.attrib["x"])
                    self.tilt = float(elem.attrib["y"])
                elif elem.tag.endswith("Zoom") and "x" in elem.attrib:
                    self.zoom = float(elem.attrib["x"])

        except Exception:
            pass

    def start_polling(self):

        self.running = True

        def loop():
            while self.running:
                self.get_status()
                time.sleep(0.35)

        self.poll_thread = threading.Thread(target=loop, daemon=True)
        self.poll_thread.start()

    def stop_polling(self):
        self.running = False
        if self.poll_thread:
            self.poll_thread.join(timeout=1)

    # =================================================
    # ABSOLUTE MOVE (PAN / TILT) - patrol only
    # =================================================

    def absolute_move(self, pan, tilt, zoom, speed=PATROL_SPEED):

        body = f"""
<tptz:AbsoluteMove
xmlns:tptz="http://www.onvif.org/ver20/ptz/wsdl"
xmlns:tt="http://www.onvif.org/ver10/schema">

<tptz:ProfileToken>{PROFILE_TOKEN}</tptz:ProfileToken>

<tptz:Position>
<tt:PanTilt x="{pan}" y="{tilt}" />
<tt:Zoom x="{zoom}" />
</tptz:Position>

<tptz:Speed>
<tt:PanTilt x="{speed}" y="{speed}" />
<tt:Zoom x="0.0" />
</tptz:Speed>

</tptz:AbsoluteMove>
"""
        self.send_soap(PTZ_URL, body)

    # =================================================
    # ZOOM-ONLY MOVE - patrol only, used while parked at HOME
    # =================================================

    def move_zoom(self, pan, tilt, zoom, zoom_speed=ZOOM_SPEED, tilt_speed=0.0):

        body = f"""
<tptz:AbsoluteMove
xmlns:tptz="http://www.onvif.org/ver20/ptz/wsdl"
xmlns:tt="http://www.onvif.org/ver10/schema">

<tptz:ProfileToken>{PROFILE_TOKEN}</tptz:ProfileToken>

<tptz:Position>
<tt:PanTilt x="{pan}" y="{tilt}" />
<tt:Zoom x="{zoom}" />
</tptz:Position>

<tptz:Speed>
<tt:PanTilt x="{tilt_speed}" y="{tilt_speed}" />
<tt:Zoom x="{zoom_speed}" />
</tptz:Speed>

</tptz:AbsoluteMove>
"""
        self.send_soap(PTZ_URL, body)

    # =================================================
    # CONTINUOUS MOVE - tracking only (visual servo)
    # =================================================

    def continuous_move(self, pan_speed, tilt_speed, zoom_speed=0.0):

        body = f"""
<tptz:ContinuousMove
xmlns:tptz="http://www.onvif.org/ver20/ptz/wsdl"
xmlns:tt="http://www.onvif.org/ver10/schema">

<tptz:ProfileToken>{PROFILE_TOKEN}</tptz:ProfileToken>

<tptz:Velocity>
<tt:PanTilt x="{pan_speed}" y="{tilt_speed}" />
<tt:Zoom x="{zoom_speed}" />
</tptz:Velocity>

</tptz:ContinuousMove>
"""
        self.send_soap(PTZ_URL, body)

    # =================================================
    # STOP (halts either AbsoluteMove or ContinuousMove)
    # =================================================

    def stop_move(self):

        body = f"""
<tptz:Stop
xmlns:tptz="http://www.onvif.org/ver20/ptz/wsdl">
<tptz:ProfileToken>{PROFILE_TOKEN}</tptz:ProfileToken>
<tptz:PanTilt>true</tptz:PanTilt>
<tptz:Zoom>true</tptz:Zoom>
</tptz:Stop>
"""
        self.send_soap(PTZ_URL, body)

    # =================================================
    # WAIT FOR ARRIVAL - interrupt-aware (patrol only)
    # =================================================
    # Waits for the camera to reach target_pan. Unlike a plain
    # "pause and resume the same move" loop, this STOPS the
    # camera immediately and returns "INTERRUPTED" the moment a
    # person is detected (shared_state.person_trigger_event set).
    # The caller (patrol_track_main.py) is responsible for
    # handling that by waiting out the tracking session and then
    # re-issuing a fresh move.

    def wait_until_reached_interruptible(
        self,
        shared_state,
        target_pan,
        tilt,
        zoom,
        speed,
        tolerance=POSITION_TOLERANCE
    ):

        stall_count = 0
        prev_pan = None

        while True:

            if shared_state.person_trigger_event.is_set():
                self.stop_move()
                return "INTERRUPTED"

            current_pan = self.pan

            if current_pan is None:
                time.sleep(LOOP_POLL_INTERVAL)
                continue

            error_direct = abs(current_pan - target_pan)
            error_wrap = 2.0 - error_direct
            error = min(error_direct, error_wrap)

            if error < tolerance:
                return "REACHED"

            if prev_pan is not None:
                if abs(current_pan - prev_pan) < 0.001:
                    stall_count += 1
                    if stall_count >= STALL_COUNT_THRESHOLD:
                        print(f"STALL detected near {current_pan:.4f}")
                        return "STALLED"
                else:
                    stall_count = 0

            prev_pan = current_pan
            time.sleep(LOOP_POLL_INTERVAL)

    # =================================================
    # GET RTSP STREAM URI (ONVIF MEDIA SERVICE) - optional,
    # not used since RTSP_URL is already known, kept for
    # completeness / fallback.
    # =================================================

    def get_stream_uri(self):

        body = f"""
<trt:GetStreamUri
xmlns:trt="http://www.onvif.org/ver10/media/wsdl"
xmlns:tt="http://www.onvif.org/ver10/schema">

<trt:StreamSetup>
<tt:Stream>RTP-Unicast</tt:Stream>
<tt:Transport>
<tt:Protocol>RTSP</tt:Protocol>
</tt:Transport>
</trt:StreamSetup>

<trt:ProfileToken>{PROFILE_TOKEN}</trt:ProfileToken>

</trt:GetStreamUri>
"""
        response = self.send_soap(MEDIA_URL, body)
        root = ET.fromstring(response.text)

        for elem in root.iter():
            if elem.tag.endswith("Uri"):
                return elem.text.strip()

        raise RuntimeError(
            "Could not retrieve RTSP stream URI from camera. "
            "Check MEDIA_URL / PROFILE_TOKEN."
        )


# =====================================================
# PTZ COMMAND THREAD - tracking only
# =====================================================
# Sends ContinuousMove / Stop at a fixed rate based on whatever
# velocity the main detection/tracking loop most recently
# computed. Only ever actually moves the camera while
# shared_state's velocity is "locked" (i.e. mode == TRACKING and
# a target is currently held) - during PATROL it stays silent so
# it never fights with the patrol thread's AbsoluteMove calls.

def ptz_command_loop(ptz, shared_state):

    period = 1.0 / COMMAND_RATE_HZ
    was_moving = False

    while True:

        pan_speed, tilt_speed, locked = shared_state.get_desired_velocity()

        moving = locked and (
            abs(pan_speed) > 0.0 or abs(tilt_speed) > 0.0
        )

        if moving:
            ptz.continuous_move(pan_speed, tilt_speed, 0.0)
            was_moving = True
        elif was_moving:
            ptz.stop_move()
            was_moving = False

        time.sleep(period)


# =====================================================
# NEW: BACKEND PPE ALERT MANAGER
# =====================================================
# Sends PPE violation alerts to the backend REST API asynchronously.
#
# KEY BEHAVIOURS:
#   - Global 15-second cooldown  (ONE alert per PPE_ALERT_COOLDOWN_SECONDS,
#     camera-wide; NOT per person, NOT per PPE item).
#   - Bounded queue (MAX_ALERT_QUEUE_SIZE=3); oldest alert is dropped
#     when the queue is full so stale data never accumulates.
#   - Single persistent background worker thread - NEVER blocks YOLO,
#     ByteTrack, PTZ, patrol, or the dashboard MJPEG feed.
#   - HTTP POST with multipart/form-data:
#       alert_data = JSON string
#       image      = JPEG snapshot (annotated frame)
#
# INTEGRATION in patrol_track_main.py:
#   alert_manager = PPEAlertManager()
#   # inside detection loop, after active_violations is populated:
#   if active_violations and alert_manager.should_send_alert():
#       _pending_ppe_alert = (target["id"], list(active_violations),
#                             dict(item_status_this_frame))
#   # after annotated_frame is ready:
#   if _pending_ppe_alert is not None:
#       pid, viols, istat = _pending_ppe_alert
#       alert_manager.queue_alert(pid, viols, istat, annotated_frame.copy())
#       _pending_ppe_alert = None
#   # on shutdown:
#   alert_manager.stop()

class PPEAlertManager:
    """
    Asynchronous PPE violation alert sender.

    The detection loop calls should_send_alert() / queue_alert().
    A single background worker thread pops from the bounded queue
    and does the HTTP POST - the detection loop never waits for it.
    """

    def __init__(self):
        self._lock            = threading.Lock()
        self._last_alert_time = 0.0  # global cooldown timestamp

        # Queue items: (alert_dict, snapshot_path_or_None)
        self._queue     = queue.Queue(maxsize=MAX_ALERT_QUEUE_SIZE)
        self._stop_evt  = threading.Event()

        # JWT token cache — managed by _get_auth_token().
        # Seeded from BACKEND_AUTH_TOKEN on startup; cleared on 401 so
        # the worker automatically refreshes or re-logs in.
        self._jwt_token     = BACKEND_AUTH_TOKEN   # pre-load static token
        self._refresh_token = BACKEND_REFRESH_TOKEN  # pre-load refresh token

        os.makedirs(BACKEND_SNAPSHOT_DIR, exist_ok=True)

        self._worker = threading.Thread(
            target=self._worker_loop,
            daemon=True,
            name="PPEAlertWorker",
        )
        self._worker.start()
        print("[PPE ALERT] Alert manager started.")
        print(f"[PPE ALERT]  Backend URL : {BACKEND_SERVER_URL}")
        print(f"[PPE ALERT]  Camera ID   : {CAMERA_ID}")
        print(f"[PPE ALERT]  Cooldown    : {PPE_ALERT_COOLDOWN_SECONDS}s")
        if BACKEND_AUTH_TOKEN:
            print("[PPE ALERT]  Auth        : Static JWT token configured.")
        elif BACKEND_USERNAME:
            print(f"[PPE ALERT]  Auth        : Auto-login as '{BACKEND_USERNAME}'.")
        else:
            print("[PPE ALERT]  Auth        : WARNING - no token or credentials set!")

    # --------------------------------------------------
    def should_send_alert(self):
        """Returns True if the global 15-second cooldown has expired.
        Thread-safe; read-only (does NOT start the cooldown)."""
        with self._lock:
            return (time.time() - self._last_alert_time) >= PPE_ALERT_COOLDOWN_SECONDS

    # --------------------------------------------------
    def queue_alert(self, person_id, violations, item_status, snapshot_frame):
        """
        Called from the detection loop (non-blocking).

        Steps:
          1. Save snapshot_frame as a JPEG to BACKEND_SNAPSHOT_DIR.
          2. Build the JSON alert payload.
          3. Mark the global cooldown as starting NOW.
          4. Put (alert_dict, snapshot_path) onto the bounded queue;
             drop the OLDEST entry first if the queue is full.

        Parameters
        ----------
        person_id     : int   - stable ByteTrack/StablePerson ID
        violations    : list  - raw class names, e.g. ["no_hardhat"]
        item_status   : dict  - {"Hardhat": "VIOLATION", "Vest": "OK", ...}
        snapshot_frame: ndarray - current annotated frame (bounding boxes)
        """
        now    = time.time()
        ts_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ts_file= datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        # Build snapshot filename: camera_01_person_5_20260820_183020.jpg
        snap_name = f"{CAMERA_ID}_person_{person_id}_{ts_file}.jpg"
        snap_path = os.path.join(BACKEND_SNAPSHOT_DIR, snap_name)

        # Save annotated snapshot to disk (detection loop, non-blocking write)
        try:
            cv2.imwrite(snap_path, snapshot_frame)
        except Exception as exc:
            print(f"[PPE ALERT] Snapshot save failed: {exc}")
            snap_path = None

        # Map raw class names (e.g. "no_hardhat") -> friendly item names
        # (e.g. "Hardhat") and deduplicate.
        friendly_violations = []
        seen = set()
        for cls_name in violations:
            mapping = PPE_CLASS_STATUS_MAP_LOWER.get(cls_name.lower())
            item_name = mapping[0] if mapping else cls_name
            if item_name not in seen:
                friendly_violations.append(item_name)
                seen.add(item_name)

        alert_dict = {
            "camera_id"  : CAMERA_ID,
            "person_id"  : person_id,
            "ppe_violated": True,
            "violations" : friendly_violations,
            "ppe_status" : item_status,       # {"Hardhat": "OK/VIOLATION", ...}
            "timestamp"  : ts_str,
        }

        # Start global cooldown NOW (before queuing avoids a race where
        # should_send_alert() returns True again before worker picks up).
        with self._lock:
            self._last_alert_time = now

        print(
            f"[PPE ALERT] Violation detected for Person #{person_id} "
            f"\u2014 Violations: {', '.join(friendly_violations)}"
        )
        print("[PPE ALERT] Queued backend alert.")

        # Non-blocking bounded put: drop oldest if full.
        if self._queue.full():
            try:
                self._queue.get_nowait()
                print("[PPE ALERT] Queue full - dropped oldest pending alert.")
            except queue.Empty:
                pass

        try:
            self._queue.put_nowait((alert_dict, snap_path))
        except queue.Full:
            pass  # extremely rare race between full-check and put - skip

    # --------------------------------------------------
    def _map_violation_type(self, violations):
        """
        Maps the list of friendly PPE item names (e.g. ["Hardhat", "Vest"])
        to the backend's (type, severity) pair, matching the exact strings
        already in the ai_alerts database:

          "no_ppe"           -> multiple items missing  -> CRITICAL
          "helmet_violation" -> Hardhat missing          -> CRITICAL
          "vest_violation"   -> Vest missing             -> MAJOR
          "gloves_violation" -> Gloves missing           -> MAJOR
          "boots_violation"  -> Boots missing            -> MAJOR
          "ppe_violation"    -> generic fallback         -> CRITICAL
        """
        if len(violations) >= 2:
            return "no_ppe", "CRITICAL"
        if not violations:
            return "ppe_violation", "CRITICAL"
        item = violations[0]
        mapping = {
            "Hardhat": ("helmet_violation", "CRITICAL"),
            "Vest"   : ("vest_violation",   "MAJOR"),
            "Gloves" : ("gloves_violation",  "MAJOR"),
            "Boots"  : ("boots_violation",   "MAJOR"),
        }
        return mapping.get(item, ("ppe_violation", "CRITICAL"))

    # --------------------------------------------------
    def _get_auth_token(self):
        """
        Returns a valid JWT access token.
        Priority:
          1. self._jwt_token  (seeded from BACKEND_AUTH_TOKEN; cleared on 401)
          2. Refresh using self._refresh_token (BACKEND_REFRESH_TOKEN)
          3. Fresh login with BACKEND_USERNAME / BACKEND_PASSWORD
        Returns empty string if all methods fail.

        NOTE: Uses self._jwt_token (mutable, cleared on expiry), NOT
        the BACKEND_AUTH_TOKEN constant directly, so that a 401 response
        actually triggers refresh/re-login instead of retrying the same
        expired static token forever.
        """
        # 1. Use cached/seeded access token
        if self._jwt_token:
            return self._jwt_token

        # 2. Try to refresh using the refresh token
        with self._lock:
            refresh_tok = self._refresh_token

        if refresh_tok:
            try:
                print("[PPE ALERT WORKER] Access token expired - refreshing ...")
                resp = requests.post(
                    BACKEND_REFRESH_URL,
                    json={"refresh": refresh_tok},
                    timeout=BACKEND_REQUEST_TIMEOUT,
                )
                if resp.status_code in (200, 201):
                    data  = resp.json()
                    token = data.get("access") or data.get("access_token")
                    if token:
                        with self._lock:
                            self._jwt_token = token
                        print("[PPE ALERT WORKER] Token refreshed successfully.")
                        return token
                    print(f"[PPE ALERT WORKER] Refresh OK but no access token in response: {data}")
                else:
                    print(f"[PPE ALERT WORKER] Token refresh failed: HTTP {resp.status_code}")
                    with self._lock:
                        self._refresh_token = ""  # refresh token also expired
            except Exception as exc:
                print(f"[PPE ALERT WORKER] Token refresh error: {exc}")

        # 3. Fall back to username/password login
        if not BACKEND_USERNAME or not BACKEND_PASSWORD:
            print("[PPE ALERT WORKER] No valid token and no login credentials set.")
            return ""

        try:
            print(f"[PPE ALERT WORKER] Logging in as '{BACKEND_USERNAME}' ...")
            resp = requests.post(
                BACKEND_LOGIN_URL,
                json={"username": BACKEND_USERNAME, "password": BACKEND_PASSWORD},
                timeout=BACKEND_REQUEST_TIMEOUT,
            )
            if resp.status_code in (200, 201):
                data = resp.json()
                tokens_obj = (data.get("data") or {}).get("tokens") or {}
                token = (
                    data.get("access")
                    or data.get("access_token")
                    or data.get("token")
                    or (data.get("data") or {}).get("access")
                    or (data.get("data") or {}).get("access_token")
                    or tokens_obj.get("access")
                )
                refresh = (
                    data.get("refresh")
                    or data.get("refresh_token")
                    or (data.get("data") or {}).get("refresh")
                    or tokens_obj.get("refresh")
                )
                if token:
                    with self._lock:
                        self._jwt_token     = token
                        self._refresh_token = refresh or ""
                    print("[PPE ALERT WORKER] Login successful, tokens cached.")
                    return token
                print(f"[PPE ALERT WORKER] Login OK but no token in response: {data}")
            else:
                print(f"[PPE ALERT WORKER] Login failed: HTTP {resp.status_code} - {resp.text[:200]}")
        except Exception as exc:
            print(f"[PPE ALERT WORKER] Login request error: {exc}")

        return ""


    # --------------------------------------------------
    def _worker_loop(self):
        """
        Background worker thread.
        Pops (alert_dict, snapshot_path) from the queue and does:

          POST /api/ai-alerts/
          Authorization: Bearer <access_token>
          Content-Type: multipart/form-data

          Fields (exact AIAlertRequest DB schema):
            type        - string  : "helmet_violation" | "vest_violation" |
                                    "gloves_violation" | "boots_violation" |
                                    "no_ppe" (2+ items missing)
            severity    - string  : "CRITICAL" | "MAJOR"
            timestamp   - string  : ISO datetime e.g. "2026-08-20 18:30:20"
            status      - string  : "OPEN"  (new alert)
            camera_id   - integer : FK camera ID in backend DB
            site_id     - integer : FK site ID in backend DB
            snapshot    - file    : JPEG annotated frame (multipart)

          Expected success: HTTP 201 Created
        """
        while not self._stop_evt.is_set():
            try:
                item = self._queue.get(timeout=1.0)
            except queue.Empty:
                continue

            alert_dict, snap_path = item
            person_id  = alert_dict.get("person_id", "?")
            violations = alert_dict.get("violations", [])

            alert_type, severity = self._map_violation_type(violations)

            print(
                f"[PPE ALERT WORKER] Sending Person #{person_id} alert "
                f"(type={alert_type}, severity={severity}) "
                f"to {BACKEND_SERVER_URL} ..."
            )

            file_handle = None
            try:
                # --- Build Authorization & Content-Type headers ---
                token   = self._get_auth_token()
                headers = {"Content-Type": "application/json"}
                if token:
                    headers["Authorization"] = f"Bearer {token}"

                # --- Encode snapshot image as base64 ---
                b64_snapshot = ""
                if snap_path and os.path.exists(snap_path):
                    try:
                        with open(snap_path, "rb") as f:
                            raw_bytes = f.read()
                            b64_snapshot = f"data:image/jpeg;base64,{base64.b64encode(raw_bytes).decode('utf-8')}"
                    except Exception as e:
                        print(f"[PPE ALERT WORKER] Error encoding snapshot image: {e}")

                # --- Build JSON payload matching Django AIAlert schema ---
                data_payload = {
                    "type"     : alert_type,
                    "severity" : severity,
                    "timestamp": alert_dict.get("timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                    "status"   : "OPEN",
                    "camera"   : int(BACKEND_CAMERA_ID),
                    "site"     : int(BACKEND_SITE_ID),
                    "snapshot" : b64_snapshot,
                }

                resp = requests.post(
                    BACKEND_SERVER_URL,
                    json    = data_payload,
                    headers = headers,
                    timeout = BACKEND_REQUEST_TIMEOUT,
                )

                if resp.status_code == 201:
                    try:
                        resp_data = resp.json()
                        alert_id  = (
                            resp_data.get("alert_id")
                            or (resp_data.get("data") or {}).get("alert_id")
                            or "?"
                        )
                    except Exception:
                        alert_id = "?"
                    print(
                        f"[PPE ALERT WORKER] Alert sent successfully "
                        f"(HTTP 201, alert_id={alert_id}) "
                        f"Person #{person_id} - {alert_type}"
                    )

                elif resp.status_code == 401:
                    # Access token expired - clear cached tokens so next
                    # attempt triggers a refresh / re-login.
                    print("[PPE ALERT WORKER] HTTP 401 Unauthorized - token expired, will refresh.")
                    with self._lock:
                        self._jwt_token    = ""
                        self._refresh_token = BACKEND_REFRESH_TOKEN  # retry with refresh

                elif resp.status_code == 400:
                    print(
                        f"[PPE ALERT WORKER] HTTP 400 Bad Request. "
                        f"Response: {resp.text[:400]}"
                    )

                else:
                    print(
                        f"[PPE ALERT WORKER] HTTP {resp.status_code} from backend. "
                        f"Response: {resp.text[:200]}"
                    )

            except requests.exceptions.ConnectionError:
                print("[PPE ALERT WORKER] Backend unavailable (connection refused)")
            except requests.exceptions.Timeout:
                print(f"[PPE ALERT WORKER] Request timeout after {BACKEND_REQUEST_TIMEOUT}s")
            except requests.exceptions.HTTPError as exc:
                print(f"[PPE ALERT WORKER] HTTP error: {exc}")
            except Exception as exc:
                print(f"[PPE ALERT WORKER] Backend request failed: {exc}")
            finally:
                if file_handle is not None:
                    try:
                        file_handle.close()
                    except Exception:
                        pass
                try:
                    self._queue.task_done()
                except Exception:
                    pass

    # --------------------------------------------------
    def stop(self):
        """Signal the worker to exit and wait (up to 3 s) for it."""
        self._stop_evt.set()
        self._worker.join(timeout=3.0)
        print("[PPE ALERT] Alert manager stopped.")